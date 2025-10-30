import random

import openpyxl

# 第一步：创建工作簿（Workbook）
wb = openpyxl.Workbook()

# 第二步：添加工作表（Worksheet）
sheet = wb.active

# 第三步：写入数据
sheet.title = '期末成绩'

# 写入标题行和学生成绩
titles = ('姓名', '语文', '数学', '英语')

# 写入标题行
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)

# 写入学生姓名和随机成绩
names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index, name in enumerate(names):
    # 写入姓名和随机成绩
    sheet.cell(row_index + 2, 1, name)
    # 写入随机成绩
    for col_index in range(2, 5):
        # 生成50到100之间的随机整数作为成绩
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# 第四步：保存工作簿
wb.save('考试成绩表.xlsx')