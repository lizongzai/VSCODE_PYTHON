"""
Version: 1.0
Author: ChatGPT
Description: 使用"考试成绩表.xlsx"调整样式和公式计算的 Excel 写入示例代码。
Date: 2025-10-30

Keyword arguments:
argument -- description
Return: return_description
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 打开已有的工作簿
workbook = openpyxl.load_workbook('考试成绩表.xlsx')

# 选择活动的工作表
sheet = workbook.active

# 设置标题行的样式
title_font = Font(bold=True, color='FFFFFF')

# 设置单元格对齐方式
alignment = Alignment(horizontal='center', vertical='center')

# 设置标题行样式
for col in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=1, column=col)
    cell.font = title_font
    cell.alignment = alignment
    cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    # 调整列宽
    sheet.column_dimensions[get_column_letter(col)].width = 15
# 计算总分并写入新列
sheet.cell(row=1, column=sheet.max_column + 1, value='总分')

# 单科成绩低于60分的标红
for row in range(2, sheet.max_row + 1):
    for col in range(2, sheet.max_column):
        cell = sheet.cell(row=row, column=col)
        if cell.value < 60:
            cell.font = Font(color='FF0000')
        else:
            cell.font = Font(color='000000')
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# 表格加上边框
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
# 应用边框和对齐方式
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = alignment

# 写入总分公式
for row in range(2, sheet.max_row + 1):
    total_formula = f"=SUM(B{row}:D{row})"
    total_cell = sheet.cell(row=row, column=sheet.max_column)
    total_cell.value = total_formula
    total_cell.alignment = alignment
# 保存工作簿到文件
workbook.save('考试成绩表_调整样式.xlsx')
