"""
Version: 1.0
Author: ChatGPT
Description: 例如在当前文件夹下有一个名为“2022年股票数据.xlsx”的 Excel 文件，
如果想读取并显示该文件的内容，可以通过如下所示的代码来完成。
Date: 2025-10-30

Keyword arguments:
argument -- description
Return: return_description
"""

import openpyxl
from datetime import datetime

# 打开Excel文件
workbook = openpyxl.load_workbook('2022年股票数据.xlsx')

# 获取所有工作表名称
print(workbook.sheetnames)

# 获取活动的工作表
sheet = workbook.worksheets[0]

# 打印工作表的维度信息
print(sheet.dimensions)

# 获取行数和列数
print(sheet.max_row, sheet.max_column)

# 获取指定单元格的值， 比如A2单元格
print(sheet['A2'].value)

# 获取指定行列的值, 比如第2行第3列
print(sheet.cell(row=2, column=3).value)

# 获取A2到C4区域的单元格
print(sheet['A2':'C4'])

# 获取第255行的所有单元格
# print(sheet['G255'].value)

# 读取所有单元的数据
# 遍历每一行
for row in range(1, sheet.max_row + 1):
    # 遍历每一列
    for col in range(1, sheet.max_column + 1):
        # print(sheet.cell(row=row, column=col).value, end='\t')
        
        # 获取单元格值
        cell = sheet.cell(row=row, column=col)
        # 获取单元格的值
        value = cell.value
        
        # 对除首行外的其他行进行数据格式化处理
        if row > 1:
            # 第1列的日期类型处理
            if col == 1 and isinstance(value, datetime):  # 判断是否为日期类型
                value = f'{value.year}年{value.month:>02d}月{value.day:>02d}日'
            # 其他列的数字类型处理成小数点后保留两位有效数字
            elif isinstance(value, (int, float)):
                # 将数字格式化为小数点后两位
                value = f'{value:.2f}'
        # 打印单元格的值
        print(value, end='\t')
    print()  # 换行
    
# 获取第5行到第10行，第2列到第4列的数据.包含表头名称
for row in sheet.iter_rows(min_row=5, max_row=10, min_col=2, max_col=4):
    for cell in row:
        print(cell.value, end='\t')
    print()  # 换行
# 获取第5行到第10行，第2列到第4列的数据.不包含表头名称
for row in sheet.iter_rows(min_row=6, max_row=10, min_col=2, max_col=4):
    for cell in row:
        print(cell.value, end='\t')
    print()  # 换行
# 获取第2列的所有数据
for col in sheet.iter_cols(min_col=2, max_col=2):
    for cell in col:
        print(cell.value)
# 获取第2列不包含表头的所有数据
for col in sheet.iter_cols(min_col=2, max_col=2, min_row=2):
    for cell in col:
        print(cell.value)
        
# 关闭Excel文件
workbook.close()