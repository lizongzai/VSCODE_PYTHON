"""
Version: 1.0
Author: ChatGPT
Description: 下面我们使用openpyxl来进行写 Excel 操作的示例代码。
Date: 2025-10-30

Keyword arguments:
argument -- description
Return: return_description
"""
import openpyxl
from openpyxl.styles import Alignment

# 创建一个新的工作簿
workbook =openpyxl.Workbook()

# 选择默认的工作表
sheet = workbook.active

# 重命名工作表
sheet.title = 'SampleSheet'

# 定义标题行
title ={'A1': 'Name', 'B1': 'Age', 'C1': 'City'}

# 写入标题行
for cell, value in title.items():
    sheet[cell] = value
    sheet[cell].alignment = alignment = Alignment(horizontal='center', vertical='center')

# 向单元格写入数据
data = [
    ['Alice', 30, 'New York'],
    ['Bob', 25, 'Los Angeles'],
    ['Charlie', 35, 'Chicago'],
    ['David', 28, 'San Francisco'],
    ['Eva', 22, 'Miami'],
    ['Frank', 33, 'Seattle'],
    ['Grace', 27, 'Boston'],
    ['Hannah', 29, 'Denver'],
    ['Ian', 31, 'Austin'],
    ['Jack', 26, 'Portland']
]

# 写入数据行

# 设置单元格对齐方式
alignment = Alignment(horizontal='center', vertical='center')

# 写入数据行
for row_index, row_data in enumerate(data, start=2):
    # 写入每一行的数据
    for col_index, cell_value in enumerate(row_data, start=1):
        # 写入单元格数据
        sheet.cell(row=row_index, column=col_index, value=cell_value)
        # 设置单元格对齐方式
        sheet.cell(row=row_index, column=col_index).alignment = alignment

# 保存工作簿到文件
workbook.save('sample_output.xlsx')
print("数据已写入 sample_output.xlsx 文件中。")
