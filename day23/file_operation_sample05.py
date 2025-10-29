"""
Version: 1.0
Author: ChatGPT
Description: 
例如在当前文件夹下有一个名为“2022年股票数据.xlsx”的 Excel 文件，
如果想读取并显示该文件的内容，可以通过如下所示的代码来完成。
Date: 2025-10-29

Keyword arguments:
argument -- description
Return: return_description
"""

import openpyxl
from datetime import datetime

# 使用openpyxl加载Excel文件
wb = openpyxl.load_workbook('2022年股票数据.xlsx')

# 获取所有工作表名称
sheetnames = wb.sheetnames
# print(sheetnames)

# 获取第一个工作表
sheet = wb[sheetnames[0]]

# 获取工作表的行数和列数
# print(sheet.max_row, sheet.max_column)

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        # 获取单元格值
        cell = sheet.cell(row, col)
        value = cell.value
        
        # 对除首行外的其他行进行数据格式化处理
        if row > 1: # 跳过表头行
            # 第1列的日期类型处理
            if col == 1 and isinstance(value, datetime): # 判断是否为日期类型
                value = f'{value.year}年{value.month:>02d}月{value.day:>02d}日'
            # 其他列的数字类型处理成小数点后保留两位有效数字
            elif isinstance(value, (int, float)): # 判断是否为数字类型
                value = f'{value:.2f}'
        print(value, end='\t')
    print()

# 获取最后一行的值（列表）
last_row_values = [sheet.cell(sheet.max_row, col).value for col in range(1, sheet.max_column + 1)]
print("最后一行:", last_row_values)


# 获取最后一行的值并格式化日期
last_row_values = [
    sheet.cell(sheet.max_row, col).value.strftime('%Y-%m-%d') # 格式化日期
    if isinstance(sheet.cell(sheet.max_row, col).value, datetime)  # 判断是否为日期类型
    else sheet.cell(sheet.max_row, col).value # 其他类型直接获取值
    for col in range(1, sheet.max_column + 1) # 获取最后一行的值
]
print("最后一行:", last_row_values)

# 获取第一行的值（列表）
first_row_values = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
print("第一行:", first_row_values)

# 获取指定行指定列范围的数据（列表）
row_data = [sheet.cell(4, col).value.strftime('%Y-%m-%d')
            if isinstance(sheet.cell(4, col).value, datetime)
            else sheet.cell(4, col).value
            for col in range(1, 6)]
print("第4行前5列:", row_data)


# 获取第5到第10行，第1到第5列的数据
rows_data = [
    [
        sheet.cell(row, col).value.strftime('%Y-%m-%d')
        if isinstance(sheet.cell(row, col).value, datetime)
        else sheet.cell(row, col).value
        for col in range(1, 6)
    ]
    for row in range(5, 11)  # 5~10行（包含5，不包含11）
]
print("5~10行数据:", rows_data)
for i, row_data in enumerate(rows_data, 5):
    print(f"第{i}行: {row_data}")