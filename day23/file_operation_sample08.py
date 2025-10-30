import openpyxl
from openpyxl.utils import get_column_letter

def calculate_stock_statistics(input_file, output_file):
    """
    统计股票数据：收盘价平均值和交易量总和
    """
    # 读取Excel文件
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    
    # 获取数据范围
    nrows = sheet.max_row
    
    # 假设列结构（根据实际文件调整）：
    # A列: 日期, B列: 开盘价, C列: 最高价, D列: 最低价, E列: 收盘价, F列: 交易量
    
    # 在数据下方添加统计行
    summary_row = nrows + 2
    
    # 添加统计标题和公式
    sheet.cell(summary_row, 1, "统计结果").font = openpyxl.styles.Font(bold=True, size=12)
    
    # 收盘价平均值（假设在第5列）
    sheet.cell(summary_row + 1, 1, "平均收盘价:")
    sheet.cell(summary_row + 1, 2, f"=AVERAGE(E2:E{nrows})")
    
    # 交易量总和（假设在第6列）
    sheet.cell(summary_row + 2, 1, "总交易量:")
    sheet.cell(summary_row + 2, 2, f"=SUM(F2:F{nrows})")
    
    # 最高收盘价
    sheet.cell(summary_row + 3, 1, "最高收盘价:")
    sheet.cell(summary_row + 3, 2, f"=MAX(E2:E{nrows})")
    
    # 最低收盘价
    sheet.cell(summary_row + 4, 1, "最低收盘价:")
    sheet.cell(summary_row + 4, 2, f"=MIN(E2:E{nrows})")
    
    # 保存文件
    workbook.save(output_file)
    print(f"统计完成！文件已保存为: {output_file}")

# 使用示例
if __name__ == "__main__":
    calculate_stock_statistics('股票数据.xlsx', '股票数据_统计.xlsx')