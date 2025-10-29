"""
Version: 1.0
Author: ChatGPT
Description: 写入Excel文件,并在写入后读取显示内容。
Date: 2025-10-29

Keyword arguments:
argument -- description
Return: return_description
"""
import openpyxl
from datetime import datetime, date
# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active  # 获取活动的工作表
# 写入表头
sheet.append(['日期', '股票代码', '股票名称', '收盘价', '涨跌幅'])
# 写入一些示例数据
data = [
    [date(2022, 1, 3), '000001', '平安银行', 15.23, 1.5],
    [date(2022, 1, 4), '000002', '万科A', 22.45, -0.8],
    [date(2022, 1, 5), '000003', '国农科技', 8.67, 2.3],
]
for row in data:
    sheet.append(row)
# 保存Excel文件
workbook.save('股票数据.xlsx')
# 读取并显示Excel文件内容
workbook = openpyxl.load_workbook('股票数据.xlsx')

# 获取活动的工作表
sheet = workbook.active

# 遍历每一行并打印
for row in sheet.iter_rows(values_only=True):
    print(row)
workbook.close()
