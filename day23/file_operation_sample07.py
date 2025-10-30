"""
Version: 2.0
Author: ChatGPT
Description: 写入Excel文件，添加丰富的样式和格式，并在写入后读取显示内容。
Date: 2025-10-29
"""

import openpyxl
from datetime import datetime, date
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, 
    NamedStyle, numbers
)
from openpyxl.formatting.rule import ColorScaleRule

def create_styles():
    """创建预定义的样式"""
    styles = {}
    
    # 表头样式
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    styles['header'] = header_style
    
    # 数据行样式 - 奇数行
    data_style_odd = NamedStyle(name="data_style_odd")
    data_style_odd.font = Font(name='宋体', size=10)
    data_style_odd.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    data_style_odd.alignment = Alignment(vertical='center')
    data_style_odd.border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    styles['data_odd'] = data_style_odd
    
    # 数据行样式 - 偶数行
    data_style_even = NamedStyle(name="data_style_even")
    data_style_even.font = Font(name='宋体', size=10)
    data_style_even.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    data_style_even.alignment = Alignment(vertical='center')
    data_style_even.border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    styles['data_even'] = data_style_even
    
    # 正涨幅样式（绿色）
    positive_style = NamedStyle(name="positive_style")
    positive_style.font = Font(name='宋体', size=10, color='00B050', bold=True)
    styles['positive'] = positive_style
    
    # 负涨幅样式（红色）
    negative_style = NamedStyle(name="negative_style")
    negative_style.font = Font(name='宋体', size=10, color='FF0000', bold=True)
    styles['negative'] = negative_style
    
    # 日期样式
    date_style = NamedStyle(name="date_style")
    date_style.font = Font(name='宋体', size=10)
    date_style.number_format = 'YYYY-MM-DD'
    date_style.alignment = Alignment(horizontal='center', vertical='center')
    styles['date'] = date_style
    
    # 价格样式
    price_style = NamedStyle(name="price_style")
    price_style.font = Font(name='Arial', size=10)
    price_style.number_format = '#,##0.00'
    price_style.alignment = Alignment(horizontal='right', vertical='center')
    styles['price'] = price_style
    
    # 涨跌幅样式
    change_style = NamedStyle(name="change_style")
    change_style.font = Font(name='Arial', size=10)
    change_style.number_format = '0.00"%"'
    change_style.alignment = Alignment(horizontal='right', vertical='center')
    styles['change'] = change_style
    
    return styles

def apply_styles_to_workbook(sheet, styles):
    """应用样式到工作表"""
    
    # 设置列宽
    sheet.column_dimensions['A'].width = 12  # 日期
    sheet.column_dimensions['B'].width = 10  # 股票代码
    sheet.column_dimensions['C'].width = 15  # 股票名称
    sheet.column_dimensions['D'].width = 12  # 收盘价
    sheet.column_dimensions['E'].width = 10  # 涨跌幅
    
    # 设置行高
    sheet.row_dimensions[1].height = 25  # 表头行高
    
    # 应用表头样式
    for col in range(1, 6):
        cell = sheet.cell(1, col)
        cell.style = styles['header']
    
    # 应用数据行样式
    for row in range(2, sheet.max_row + 1):
        for col in range(1, 6):
            cell = sheet.cell(row, col)
            
            # 交替行颜色
            if row % 2 == 0:
                cell.style = styles['data_even']
            else:
                cell.style = styles['data_odd']
            
            # 特定列的样式
            if col == 1:  # 日期列
                cell.style = styles['date']
            elif col == 4:  # 收盘价列
                cell.style = styles['price']
            elif col == 5:  # 涨跌幅列
                cell.style = styles['change']
                # 根据涨跌幅值设置颜色
                if cell.value is not None:
                    if cell.value > 0:
                        cell.font = styles['positive'].font
                    elif cell.value < 0:
                        cell.font = styles['negative'].font

def add_conditional_formatting(sheet):
    """添加条件格式"""
    
    # 涨跌幅颜色刻度规则
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='FF0000',  # 红色（最低）
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',  # 白色（中间）
        end_type='max', end_color='00FF00'  # 绿色（最高）
    )
    sheet.conditional_formatting.add(f'E2:E{sheet.max_row}', color_scale_rule)

def main():
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()
    # 获取活动的工作表
    sheet = workbook.active
    # 设置工作表标题
    sheet.title = "股票数据"
    
    # 创建样式
    styles = create_styles()
    
    # 注册命名样式
    for style_name, style_obj in styles.items():
        # 仅在样式未注册时添加
        if style_name not in workbook.named_styles:
            workbook.add_named_style(style_obj)
    
    # 写入表头
    headers = ['日期', '股票代码', '股票名称', '收盘价', '涨跌幅(%)']
    sheet.append(headers)
    
    # 写入示例数据
    data = [
        [date(2022, 1, 3), '000001', '平安银行', 15.23, 1.5],
        [date(2022, 1, 4), '000002', '万科A', 22.45, -0.8],
        [date(2022, 1, 5), '000003', '国农科技', 8.67, 2.3],
        [date(2022, 1, 6), '000004', '国华网安', 12.34, 0.5],
        [date(2022, 1, 7), '000005', '世纪星源', 5.67, -1.2],
        [date(2022, 1, 8), '000006', '深振业A', 18.90, 3.1],
        [date(2022, 1, 9), '000007', '全新好', 9.87, -2.5],
        [date(2022, 1, 10), '000008', '神州高铁', 4.56, 0.9],
    ]
    
    for row in data:
        sheet.append(row)
    
    # 应用样式
    apply_styles_to_workbook(sheet, styles)
    
    # 添加条件格式
    add_conditional_formatting(sheet)
    
    # 添加冻结窗格（冻结表头）
    sheet.freeze_panes = 'A2'
    
    # 保存Excel文件
    workbook.save('股票数据_带样式.xlsx')
    print("带样式的Excel文件已生成: 股票数据_带样式.xlsx")
    
    # 读取并显示Excel文件内容
    print("\n读取Excel文件内容:")
    print("=" * 80)
    
    workbook_read = openpyxl.load_workbook('股票数据_带样式.xlsx')
    sheet_read = workbook_read.active
    
    # 遍历每一行并打印
    for row_num, row in enumerate(sheet_read.iter_rows(values_only=True), 1):
        if row_num == 1:
            # 表头加粗显示
            print(f"\033[1m{row}\033[0m")  # 使用ANSI转义序列加粗
        else:
            formatted_row = list(row)
            # 格式化显示
            formatted_row[3] = f"{formatted_row[3]:.2f}" if formatted_row[3] else "N/A"
            formatted_row[4] = f"{formatted_row[4]:+.2f}%" if formatted_row[4] is not None else "N/A"
            print(formatted_row)
    
    workbook_read.close()
    print("=" * 80)

if __name__ == "__main__":
    main()